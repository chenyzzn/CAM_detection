import os
import pandas as pd
import openpyxl
import datetime
import cv2
import shutil
from os import listdir
from os.path import isfile, join
import sqlite3

def folder_file_num(path_dir):
    source = os.walk(path_dir)
    i_box=0
    j_item=0
    for folder, subfolders, filenames in source:
        print(f'目前資料夾路徑為：{folder}')
        for subfolder in subfolders:
            j_item=j_item+1
            print(f'{folder}的子資料夾為：{subfolder}')
        for filename in filenames:
            i_box=i_box+1
            #print(i_box," ___", f'{folder}/{subfolder}內含檔案為：{filename}')
    print("品項數= " + str(j_item) , " ,箱數= " + str(i_box) ) 
    return i_box,j_item


print("參數收集.....")
dfdate = pd.read_excel("D:\\PartData\\px_main.xlsx") #, dtype = {"ITF":"object", '貨號':'object'}
dfdate['ITF'] = dfdate['ITF'].apply(lambda x : '{:0>14d}'.format(x))
dfdate['貨號'] = dfdate['貨號'].apply(lambda x : '{:0>8d}'.format(x))
dfdate = dfdate.set_index("ITF")
wb = openpyxl.load_workbook('D:\\PartData\\ocr_record.xlsx') #報表模板

today = datetime.datetime.now().strftime("%Y-%m-%d") #string

conn = sqlite3.connect('D:\\Partdata\\NG.db')
cur = conn.cursor()
df = pd.read_sql_query('SELECT * FROM NG;', conn)

#NG_file = open(f'D:\\partData\\NG_file\\NG_file_{today}.txt', 'r+')
#NG_file_TCAM = open(f'D:\\partData\\NG_file_TCAM\\NG_file_{today}.txt', 'r+')

print("參數收集完成")

#NG_list = NG_file.readlines()
#for i in range(len(NG_list)):
#    NG_list[i] = NG_list[i].split()

#df = pd.DataFrame(columns = ['filename', 'expd'], data = NG_list)

#df['ITF'] = df['filename'].str[34:48]
r_NG = df[df['CAM'] == 'R_Cam'].groupby(['ITF', 'expd']).size().reset_index(name = 'count')
t_NG = df[df["CAM"] == "T_Cam"].groupby(['ITF', 'expd']).size().reset_index(name = 'count')
#print(r_NG)
'''
NG_list_TCAM = NG_file_TCAM.readlines()
for i in range(len(NG_list_TCAM)):
    NG_list_TCAM[i] = NG_list_TCAM[i].split()

df2 = pd.DataFrame(columns = ['filename', 'expd'], data = NG_list_TCAM)
df2['ITF'] = df2['filename'].str[34:48]
grouped_NG2 = df2.groupby(['ITF', 'expd']).agg('count').reset_index()
'''
'''
path1="D:\\OCR_EXP_NG\\"
path2="D:\\OCR_DATA_NG\\"
list1 = os.listdir(path1)
#如果沒目錄就造一個 
if not os.path.isdir(path2): 
    os.mkdir(path2)
#shutil.rmtree(path2)

for filename in list1:
    img0=cv2.imread(path1 + filename)
    num = filename.find(".jpg") #長度
    itemname = filename[num-16:num-2]#貨號 
    if not os.path.isdir(path2):  os.mkdir(path2)#如果沒目錄就造一個 
    path3=path2+itemname
    if not os.path.isdir(path3):  os.mkdir(path3)#如果沒目錄就造一個
    cv2.imwrite(path3 + "\\" + filename ,img0)

#today = datetime.now().strftime("%Y-%m-%d")
#if os.path.isfile(f"D:\\OCR_DATA_NG\\ocr_record_{today}.xlsx")==True:
#os.remove(f"D:\\OCR_DATA_NG\\ocr_record_{today}.xlsx")
'''

#wb.create_sheet("效期辨識", 0)  插入工作表 0 在第一個位置
s1 = wb['RCAM']
s2 = wb['TCAM']

#===============NG資料=========================================
i=16    
j = 0
for name in r_NG['ITF']: 
        expd = r_NG.iloc[j]['expd']
        y = int(expd[0:4])
        m = int(expd[4:6])
        d = int(expd[6:8])
        expdate=datetime.date(y,m,d)
        datecheck= datetime.date(y,m,d) - datetime.date.today() #與今天相差天數(datetime)
        checkday=int(datecheck.days) #與今天相差天數(int)
        if checkday < 0: #expd是製造日期
            useday = int(dfdate.loc[name]["保存天數"]) #保存天數
            checkday = int(datecheck.days) + useday 
            expdate = datetime.date(y,m,d) + datetime.timedelta(days=useday)

        st=dfdate.loc[name]["品名"]
        stritem =str(dfdate.loc[name]["貨號"])#貨號
        
        fn = r_NG.iloc[j]['count']
        
        s1.cell(i,1).value= stritem #貨號
        s1.cell(i,2).value= name #條碼（ITF）
        s1.cell(i,3).value= st #品名
        s1.cell(i,4).value= fn #箱數
        s1.cell(i,5).value= expdate #效期
        s1.cell(i,6).value= datetime.date.today() #進貨日期
        s1.cell(i,7).value= checkday #可用天數
        s1.cell(i,8).value= dfdate.loc[name]["保存天數"]#保存天數
        s1.cell(i,9).value= dfdate.loc[name]["允收天數"] #允收天數
        s1.cell(i,10).value= dfdate.loc[name]["允出天數"] #允出天數
        print(st)
        i+=1
        j+=1

i=16    
j = 0
for name in t_NG['ITF']: 
        expd = t_NG.iloc[j]['expd']
        y = int(expd[0:4])
        m = int(expd[4:6])
        d = int(expd[6:8])
        expdate=datetime.date(y,m,d)
        datecheck= datetime.date(y,m,d) - datetime.date.today() #與今天相差天數(datetime)
        checkday=int(datecheck.days) #與今天相差天數(int)
        if checkday < 0: #expd是製造日期
            useday = int(dfdate.loc[name]["保存天數"]) #保存天數
            checkday = int(datecheck.days) + useday 
            expdate = datetime.date(y,m,d) + datetime.timedelta(days=useday)

        st=dfdate.loc[name]["品名"]
        stritem =str(dfdate.loc[name]["貨號"])#貨號
        
        fn = t_NG.iloc[j]['count']
        
        s2.cell(i,1).value= stritem #貨號
        s2.cell(i,2).value= name #條碼（ITF）
        s2.cell(i,3).value= st #品名
        s2.cell(i,4).value= fn #箱數
        s2.cell(i,5).value= expdate #效期
        s2.cell(i,6).value= datetime.date.today() #進貨日期
        s2.cell(i,7).value= checkday #可用天數
        s2.cell(i,8).value= dfdate.loc[name]["保存天數"]#保存天數
        s2.cell(i,9).value= dfdate.loc[name]["允收天數"] #允收天數
        s2.cell(i,10).value= dfdate.loc[name]["允出天數"] #允出天數
        print(st)
        i+=1
        j+=1

#=============當日統計數據=======================================        
p1 = "E:\\OCR_ERROR\\"
p2 = "E:\\OCR_ERROR_TXT\\"
p3 = "E:\\OCR_EXP\\"
p4 = "E:\\OCR_EXP_NG\\"
p5 = "E:\\OCR_flow_box\\"
p6 = "E:\\OCR_NoRead_box"
p7 = "E:\\OCR_SBS\\"

c1 = len([f for f in listdir(p1) if isfile(join(p1, f))])
#c2 = len([f for f in listdir(p2) if isfile(join(p2, f))])
ic2_box,jc2_item=folder_file_num(p2)
#c3 = len([f for f in listdir(p3) if isfile(join(p3, f))])
i_box,j_item = folder_file_num(p3)
i4_box, j_4item = folder_file_num(p4)
#c4 = len([f for f in listdir(p4) if isfile(join(p4, f))])
c5 = len([f for f in listdir(p5) if isfile(join(p5, f))])
c6 = len([f for f in listdir(p6) if isfile(join(p6, f))])
c7 = len([f for f in listdir(p7) if isfile(join(p7, f))])

s1['C1'].value = today #今天日期

s1['B5'].value = c1+ic2_box+i_box+i4_box +c6+c5+c7 #R_CAM總拍照箱數
s1['B6'].value = c1 #OCR_ERROR
s1['B7'].value = ic2_box #OCR_ERROR_TXT
s1['B8'].value = c5 #OCR_flow_box
s1['B9'].value = c6 #NoRead
s1['B10'].value = c7 #併箱

s1['D6'].value = i_box #辨識成功符合效期規範(OCR_EXP)
s1['D7'].value = i4_box #辨識成功未符合效期規範(OCR_EXP_NG)
list_exp=len(os.listdir(p3))



p1 = "F:\\OCR_ERROR\\"
p2 = "F:\\OCR_ERROR_TXT\\"
p3 = "F:\\OCR_EXP\\"
p4 = "F:\\OCR_EXP_NG\\"
p5 = "F:\\OCR_flow_box\\"
p6 = "F:\\OCR_NoRead_box"
p7 = "F:\\OCR_SBS\\"

c1 = len([f for f in listdir(p1) if isfile(join(p1, f))])
#c2 = len([f for f in listdir(p2) if isfile(join(p2, f))])
ic2_box,jc2_item=folder_file_num(p2)
#c3 = len([f for f in listdir(p3) if isfile(join(p3, f))])
i_box,j_item = folder_file_num(p3)
i4_box, j_4item = folder_file_num(p4)
#c4 = len([f for f in listdir(p4) if isfile(join(p4, f))])
c5 = len([f for f in listdir(p5) if isfile(join(p5, f))])
c6 = len([f for f in listdir(p6) if isfile(join(p6, f))])
c7 = len([f for f in listdir(p7) if isfile(join(p7, f))])


s2['C1'].value = today #今天日期

s2['B5'].value = c1+ic2_box+i_box+i4_box +c6+c5+c7 #R_CAM總拍照箱數
s2['B6'].value = c1 #OCR_ERROR
s2['B7'].value = ic2_box #OCR_ERROR_TXT
s2['B8'].value = c5 #OCR_flow_box
s2['B9'].value = c6 #NoRead
s2['B10'].value = c7 #併箱

s2['D6'].value = i_box #辨識成功符合效期規範(OCR_EXP)
s2['D7'].value = i4_box #辨識成功未符合效期規範(OCR_EXP_NG)
list_exp=len(os.listdir(p3))

try:
    if os.path.isfile(f"D:\\report\\ocr_record_{today}.xlsx")==True:
        os.remove(f"D:\\report\\ocr_record_{today}.xlsx")
    wb.save(f"D:\\report\\ocr_record_{today}.xlsx")
    print(f"D:\\report\\ocr_record_{today}.xlsx")
except:
    print("請關閉EXCEL")

#input_data = input('按Enter鍵退出:')