import os
import pandas as pd
import openpyxl
import datetime
import shutil
import sqlite3
from PyQt5 import QtWidgets
import threading
from PyQt5.QtWidgets import QApplication, QMainWindow
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import sys, demo_ui

class Worker(QObject):
    finished = pyqtSignal()  # 任務完成信號
    progress = pyqtSignal(int)  # 進度信號

    def __init__(self, handler):
        super().__init__()
        self.handler = handler
        self.run = True
        self.paused = False

    def running(self):
        while self.run:
            while self.paused:
                QThread.msleep(100) 
            self.handler.run()
            self.finished.emit()
            break

    def stop(self):
        self.run = False

    def pause(self):
        self.paused = True

    def resume(self):
        self.paused = False

class MainWindow(QMainWindow, demo_ui.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton_3.clicked.connect(self.start_task)
        self.pushButton_4.clicked.connect(self.pause_task)
        self.pushButton_5.clicked.connect(self.stop_task)

        self.thread = None
        self.worker = None
        self.paused = False

    def start_task(self):
        if not self.thread or not self.thread.isRunning():
            self.handler = Handler()
            self.thread = QThread()
            self.worker = Worker(self.handler)
            self.worker.moveToThread(self.thread)

            self.thread.started.connect(self.worker.running)
            self.worker.finished.connect(self.thread.quit)
            self.worker.finished.connect(self.worker.deleteLater)
            self.thread.finished.connect(self.thread.deleteLater)

            self.thread.start()
            print("Task started.")
        else:
            print("Task is already running.")

    def stop_task(self):
        if self.worker:
            self.worker.run = False
            print("Task stopped.")

    def pause_task(self):
        if self.worker:
            if not self.paused:
                self.worker.pause()
                self.paused = True
                self.pushButton_4.setText("Resume")
                print("Task paused.")
            else:
                self.worker.resume()
                self.paused = False
                self.pushButton_4.setText("Pause")
                print("Task resumed.")

class Handler():
    def run(self):
        self.folder_file_num()
        self.query()
        self.task()
    
    def folder_file_num(self, path_dir):
        source = os.walk(path_dir)
        i_box=0
        j_item=0
        for folder, subfolders, filenames in source:
            print(f'目前資料夾路徑為：{folder}')
            for subfolder in subfolders:
                j_item=j_item+1
                print(f'{folder}的子資料夾為：{subfolder}')
            for filename in filenames:
                i_box=i_box+1
                #print(i_box," ___", f'{folder}/{subfolder}內含檔案為：{filename}')
        print("品項數= " + str(j_item) , " ,箱數= " + str(i_box) ) 
        return i_box,j_item

        print("參數收集.....")
        #載入允收允出主檔
        dfdate = pd.read_excel("C:\\AI_exp\\HP\\Partdata\\px_main.xlsx")
        dfdate['ITF'] = dfdate['ITF'].apply(lambda x : '{:0>14}'.format(x))
        dfdate['貨號'] = dfdate['貨號'].apply(lambda x : '{:0>8}'.format(x))
        dfdate = dfdate.set_index("ITF")
        #載入短允收主檔
        short_exp = pd.read_excel("C:\\AI_exp\\HP\\Partdata\\short_exp.xlsx")
        short_exp['ITF'] = short_exp['ITF'].apply(lambda x : '{:0>14}'.format(x))
        short_exp['開始異動日期'] = short_exp['開始異動日期'].apply(lambda x : datetime.datetime.strptime(x, '%Y/%m/%d'))
        short_exp['截止日期'] = short_exp['截止日期'].apply(lambda x : datetime.datetime.strptime(x, '%Y/%m/%d'))
        short_exp = short_exp[(short_exp['開始異動日期']<=datetime.datetime.now())&(short_exp['截止日期']>=datetime.datetime.now())]
        short_itf = short_exp['ITF'].unique()
        #載入報表模板
        wb = openpyxl.load_workbook('C:\\AI_exp\\HP\\Partdata\\ocr_record.xlsx') #報表模板

        time = datetime.datetime.now().strftime("%H%M")
        time_int = int(time)
        today = datetime.datetime.now().strftime("%Y-%m-%d")

        if time_int < 300:
            yday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
            day = (today, yday)
        elif time_int < 600:
            yday = (datetime.datetime.now() - datetime.timedelta(days=1)).strftime("%Y-%m-%d")
            stime = "00:00:00" 
            etime = "03:00:00"
            day = (today, yday)
        else:
            day = (today, today)

        #載入NG列表
        conn = sqlite3.connect('C:\\AI_exp\\HP\\Partdata\\OCR2.db')
        cur = conn.cursor()
        if time_int >= 300 and time_int < 600:
            query = """SELECT * FROM RCAM WHERE (date = ?) OR (date = ? AND time BETWEEN ? AND ?)"""
            r_NG = pd.read_sql_query(query, conn, params=(yday, today, stime, etime))
            query = """SELECT * FROM TCAM WHERE (date = ?) OR (date = ? AND time BETWEEN ? AND ?)"""
            t_NG = pd.read_sql_query(query, conn, params=(yday, today, stime, etime))
        else:
            query = "SELECT * FROM RCAM WHERE date IN (?, ?) AND status = 'NG';"
            r_NG = pd.read_sql_query(query, conn, params=day)
            query = "SELECT * FROM TCAM WHERE date IN (?, ?) AND status = 'NG';"
            t_NG = pd.read_sql_query(query, conn, params=day)
        #r_NG = pd.read_sql_query(f"SELECT * FROM RCAM WHERE date = '{today}' AND status = 'NG';", conn)
        #t_NG = pd.read_sql_query(f"SELECT * FROM TCAM WHERE date = '{today}' AND status = 'NG';", conn)
        conn.close()

        #載入當天分揀指示
        try:
            conn = sqlite3.connect('C:\\AI_exp\\HP\\sorter\\px.sqlite')
            query_sorter = '''WITH cte AS (
            SELECT *, CASE WHEN LENGTH(IJP) = 8 THEN '寄庫' WHEN LENGTH(IJP) = 3 THEN '越庫' END AS xd_dc , count(*)
            FROM sorter_job
            GROUP BY BARCODE1, xd_dc), 

            cte2 AS (SELECT BARCODE1, xd_dc, count(*) as cn FROM cte
            GROUP BY BARCODE1)

            SELECT *, case when cn = 1 then 'OK' else '撞單' END AS check_ FROM cte2;'''

            sorter = pd.read_sql_query(query_sorter, conn)
            sorter = sorter.set_index("BARCODE1")
            conn.close()
        except:pass

        print("參數收集完成")

    #查詢數量
    def query(self, CAM:str, date:str, status:str) -> int:
        conn = sqlite3.connect('C:\\AI_exp\\HP\\Partdata\\OCR2.db')
        cur = conn.cursor()
        
        query = f'''SELECT count(*) FROM {CAM} WHERE  date = '{date}' AND status = '{status}';'''
        
        res = cur.execute(query)
        result = res.fetchall()[0][0]
        return result

        try: #清空OCR_EXP_NG2
            shutil.rmtree('D:\\R_CAM\\OCR_EXP_NG2\\'+today)
            #shutil.rmtree('F:\\OCR_EXP_NG2\\'+today)
        except: pass


    def task(self, CAM, today) -> None: # CAM = 'RCAM' or 'TCAM'
        choose_wb = {'RCAM':wb['RCAM'], 'TCAM':wb['TCAM']}
        choose_NG_list = {'RCAM':r_NG, 'TCAM':t_NG}
        choose_Cam = {'RCAM':'R_Cam', 'TCAM':'T_Cam'}
        s = choose_wb.get(CAM) 
        NG_list = choose_NG_list.get(CAM)
        Cam = choose_Cam.get(CAM)
        
        NG_list['mfgdate'] = ''
        NG_list['checkday'] = ''
        NG_list['result'] = ''
        NG_list['overday'] = ''


        j = 0
        
        #無超出允收0，超出允收1，超出允出2
        for j in range(NG_list.shape[0]):
            itf = NG_list.iloc[j]['itemname']
            expd = NG_list.iloc[j]['expd']
            
            y = int(expd[0:4])
            m = int(expd[4:6])
            d = int(expd[6:8])
            expdate=datetime.date(y,m,d)
            datecheck= datetime.date(y,m,d) - datetime.date.today()+datetime.timedelta(days = 1) #與今天相差天數(datetime)
            checkday=int(datecheck.days) #與今天相差天數(int)
            mfgdate = ''
            if expdate <= datetime.date.today(): #expd是製造日期
                mfgdate = expdate
                useday = int(dfdate.loc[itf]["保存天數"]) #保存天數
                checkday = int(datecheck.days) + useday
                expdate = datetime.date(y,m,d) + datetime.timedelta(days=useday)
            
            NG_list.loc[j, ['mfgdate']] = mfgdate
            NG_list.loc[j, ['checkday']] = checkday
            NG_list.loc[j, ['expd']] = expdate
            
            
            try:
                if sorter.loc[itf]['check_']=='OK':
                    xd_dc = sorter.loc[itf]['xd_dc']
                else: xd_dc = '撞單'

                if xd_dc == '越庫':
                    overday = checkday - dfdate.loc[itf]["允收天數"]
                    result = [0 if overday >= 0 else 1][0]
                elif xd_dc == '寄庫':
                    overday = checkday - round(dfdate.loc[itf]["允出天數"],0)
                    result = [0 if overday >= 0 else 2][0]
                elif xd_dc == '撞單':
                    overday = checkday - dfdate.loc[itf]["允收天數"]
                    result = [0 if overday >= 0 else 1][0]
                else:
                    xd_dc = ''
                    result = ''
                    overday = '' 
            except:
                xd_dc = ''
                result = ''
                overday = ''      
            
            if itf in short_itf: result = 0
            if overday == '' or overday < -110 : result=0
            NG_list.loc[j, ['result']] = result
            NG_list.loc[j, ['overday']] = overday
            j+=1


        #更新NG_list，排除正常貨品
        NG_list = NG_list[NG_list['result']!=0]
        NG_list = NG_list.reset_index()
        #依照ITF分組，並計算數量
        NG_group = NG_list.groupby(['itemname']).agg({'date':'count'}).reset_index().rename(columns={'date':'count'})
        
        #如果某ITF貨品NG只有1張，其他同ITF的有辨識出來且大於兩張，就剔除
        for k, itemname in enumerate(NG_group[NG_group['count'] == 1]['itemname']):
            conn = sqlite3.connect('C:\\AI_exp\\HP\\Partdata\\OCR2.db')
            cur = conn.cursor()
            res = cur.execute(f"SELECT count(*) FROM {CAM} WHERE date = '{today}' AND itemname = '{itemname}' AND status= 'EXP_OK';").fetchall()[0][0]
            conn.close()
            if res>=2:
                NG_list.loc[NG_list['itemname']==itemname, 'result'] = 0

        #重新排除並分組
        NG_list = NG_list[NG_list['result']!=0]
        NG_list = NG_list.reset_index()
        NG_group = NG_list.groupby(['itemname', 'expd', 'mfgdate', 'result','checkday', 'overday']).agg({'date':'count'}).reset_index().rename(columns={'date':'count'})

        i=16  
        j=0  
        #寫上excel
        for j in range(NG_group.shape[0]):
            
            itf = NG_group.iloc[j]['itemname']
            s.cell(i,1).value= dfdate.loc[itf]['貨號'] #貨號
            s.cell(i,2).value= itf #條碼（ITF）
            s.cell(i,3).value= dfdate.loc[itf]['品名'] #品名
            s.cell(i,4).value= NG_group.iloc[j]['count'] #箱數
            s.cell(i,5).value= NG_group.iloc[j]['mfgdate'] #製造日期
            s.cell(i,6).value= NG_group.iloc[j]['expd'] #效期
            s.cell(i,7).value= datetime.date.today() #進貨日期
            s.cell(i,8).value= NG_group.iloc[j]['checkday'] #可用天數
            s.cell(i,9).value= dfdate.loc[itf]["保存天數"]#保存天數
            s.cell(i,10).value= dfdate.loc[itf]["允收天數"] #允收天數
            s.cell(i,11).value= round(dfdate.loc[itf]["允出天數"], 0) #允出天數
            s.cell(i,12).value = ['越庫' if NG_group.iloc[j]['result']==1 else '寄庫'][0] #貨品類型
            s.cell(i,13).value = NG_group.iloc[j]['result'] #判斷結果
            s.cell(i,14).value = NG_group.iloc[j]['overday'] #超出天數   
            i+=1     
            print(itf, dfdate.loc[itf]['品名'])
            j+=1

        #存圖片
        j=0
        for j, filename in enumerate(NG_list['filename']):
            #dirt = ['E:\\' if CAM == 'RCAM' else 'F:\\'][0]
            dirt = ['D:\\'][0]
            itf = NG_list.loc[j]['itemname']
            try:
                filepath = os.path.join(dirt, 'OCR_EXP_NG',today,itf,filename)
                new_filepath = os.path.join(dirt, 'OCR_EXP_NG2',today,itf,filename)
                if not os.path.isdir(os.path.join(dirt, 'OCR_EXP_NG2',today)):  os.mkdir(os.path.join(dirt, 'OCR_EXP_NG2',today)) 
                if not os.path.isdir(os.path.join(dirt, 'OCR_EXP_NG2',today,itf)):  os.mkdir(os.path.join(dirt, 'OCR_EXP_NG2',today,itf)) 
                shutil.copy(filepath, new_filepath)
            except:pass
            j+=1

        #查詢統計數量
        flowbox = query(CAM, today, 'FL')
        sbs = query(CAM, today, 'SBS')
        noread = query(CAM, today, 'NOREAD')
        error = query(CAM, today, 'ERROR')
        error_txt = query(CAM, today, 'ERROR_TXT')
        exp = query(CAM, today, 'EXP_OK')
        ng = query(CAM, today, 'NG')

        #寫上統計數量
        s['C1'].value = str(day)
        if time_int < 300:
            s['C2'].value = f'06:00-{time}'
        elif time_int < 600:
            s['C2'].value = '06:00-03:00'
        else:
            s['C2'].value = f'06:00-{time}'

        s['B5'].value = error+error_txt+flowbox+noread+sbs+exp+ng #R_CAM總拍照箱數

        s['B6'].value = error #OCR_ERROR
        s['B7'].value = error_txt #OCR_ERROR_TXT
        s['B8'].value = flowbox #OCR_flow_box
        s['B9'].value = noread #NoRead
        s['B10'].value = sbs #併箱

        #NG數量
        xd = NG_list[NG_list['result'] == 1].shape[0] #不符允收數
        dc = NG_list[NG_list['result'] == 2].shape[0] #不符允出數
        exp = exp+(ng-xd-dc)
        s['D6'].value = exp #辨識成功符合效期規範(OCR_EXP)
        s['D7'].value = xd #辨識成功未符合允收規範
        s['D8'].value = dc #辨識成功未符合允出規範

        task('RCAM', today=today)
        task('TCAM', today=today)

        #存檔=========================================================
        try:
            if os.path.isfile(f"C:\\AI_exp\\HP\\report\\ocr_record_{today}.xlsx")==True:
                os.remove(f"C:\\AI_exp\\HP\\report\\ocr_record_{today}.xlsx")
            wb.save(f"C:\\AI_exp\\HP\\report\\ocr_record_{today}.xlsx")
            print(f"C:\\AI_exp\\HP\\report\\ocr_record_{today}.xlsx")
        except:
            print("請關閉EXCEL")  
 
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


